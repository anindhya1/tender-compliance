import json
from openpyxl import load_workbook

# --- Configuration ---
FILE_PATH = "data/checklist/bqc-checklist.xlsx"  # Your Excel file
SHEET_NAME = "Sheet1"         # Name of the tab (usually 'Sheet1')

# Column Indices (1-based: A=1, B=2, C=3...)
ID_COL = 1    # Column A contains the ID (e.g., "REQ-001")
REQ_COL = 2   # Column B contains the Requirement text

def extract_rubric_to_dict(file_path):
    # 1. Load the Excel File (data_only=True reads values, not formulas)
    wb = load_workbook(filename=file_path, data_only=True)
    
    # 2. Select the specific sheet
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active  # Fallback to active sheet
        
    rubric_dict = {}

    print(f"Reading '{file_path}'...")

    # 3. Iterate through rows (min_row=2 skips the Header)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # extract values based on configured column indices
        # Note: python lists are 0-indexed, so we subtract 1
        req_id = row[ID_COL - 1]
        req_text = row[REQ_COL - 1]

        # Only store if both ID and Text exist (skips empty rows)
        if req_id and req_text:
            rubric_dict[req_id] = req_text

    return rubric_dict

# --- Execution ---
if __name__ == "__main__":
    # Run the function
    my_rubric = extract_rubric_to_dict(FILE_PATH)

    # Print the result nicely
    print("\n--- Extracted Rubric (Dictionary) ---")
    print(json.dumps(my_rubric, indent=4))

    # Verification count
    print(f"\nTotal Requirements Found: {len(my_rubric)}")
