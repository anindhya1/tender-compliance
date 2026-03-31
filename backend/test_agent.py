import os
import json
import shutil
from pathlib import Path
from openpyxl import load_workbook

from llama_index.core import (
    VectorStoreIndex,
    SimpleDirectoryReader,
    StorageContext,
    load_index_from_storage,
    Settings,
)
from llama_index.llms.ollama import Ollama
from llama_index.embeddings.ollama import OllamaEmbedding
from llama_index.core.tools import QueryEngineTool
from llama_index.core.agent import ReActAgent

import read_checklist as checklist

# --- CONFIGURATION ---
CHECKLIST_PATH = "data/checklist/bqc-checklist.xlsx"
EVIDENCE_ROOT = "./data/markdown/01_md/bidder-docs"
TENDER_DOCS_ROOT = "./data/markdown/01_md/tender-docs"
OUTPUT_DIR = "./data/reports"
INDEX_DIR = "./data/agent_indices"

# AI Settings
LLM_MODEL = "mistral"
EMBED_MODEL = "nomic-embed-text"
LLM_TEMPERATURE = 0.1

# Excel Column Definitions
COL_ID = 1; COL_REQ = 2; COL_STATUS = 3; COL_EVIDENCE = 4

# --- LLAMA INDEX SETTINGS ---
Settings.llm = Ollama(model=LLM_MODEL, temperature=LLM_TEMPERATURE, request_timeout=1000.0)
Settings.embed_model = OllamaEmbedding(model_name=EMBED_MODEL)


# --- INDEX BUILDER ---
def build_or_load_index(name: str, docs_path: str) -> VectorStoreIndex:
    """Build a VectorStoreIndex from a folder, or load from disk if already built."""
    persist_dir = os.path.join(INDEX_DIR, name)

    if os.path.exists(persist_dir):
        print(f"   Loading existing index: '{name}'...")
        storage_context = StorageContext.from_defaults(persist_dir=persist_dir)
        return load_index_from_storage(storage_context)

    print(f"   Building index: '{name}'...")
    documents = SimpleDirectoryReader(docs_path, recursive=True).load_data()
    index = VectorStoreIndex.from_documents(documents)
    index.storage_context.persist(persist_dir=persist_dir)
    print(f"   Index saved to '{persist_dir}'.")
    return index


# --- AGENT AUDITOR ---
def audit_with_agent(agent: ReActAgent, req_id: str, req_text: str) -> dict:
    """Run the ReActAgent to evaluate a single requirement."""
    prompt = f"""
    You are a strict Compliance Auditor. Your task is to evaluate whether a bidder satisfies a requirement.

    REQUIREMENT ({req_id}): "{req_text}"

    INSTRUCTIONS:
    1. Use the tender_context tool to understand what this requirement means.
    2. Use the bidder_docs tool to find evidence of whether the bidder satisfies it.
    3. Based on your findings, determine the status:
       - "Pass" if the bidder's documents satisfy the requirement
       - "Fail" if the bidder's documents contradict or fail the requirement
       - "N/A" if no relevant evidence is found in the bidder's documents
    4. Respond ONLY with valid JSON in this exact format:
    {{
        "status": "Pass" | "Fail" | "N/A",
        "reasoning": "brief explanation",
        "quote": "exact text from bidder documents"
    }}
    """

    try:
        response = agent.chat(prompt)
        raw = str(response).strip()

        print(f"\n[AGENT RESPONSE - {req_id}]\n{raw}\n[END]\n")

        # JSON Cleanup
        if "```" in raw:
            start = raw.find("{", raw.find("```"))
            end = raw.rfind("}") + 1
            raw = raw[start:end]
        elif "{" in raw:
            start = raw.find("{")
            end = raw.rfind("}") + 1
            raw = raw[start:end]

        return json.loads(raw)

    except Exception as e:
        print(f"   [Error] Agent failed for {req_id}: {e}")
        return {"status": "Error", "reasoning": str(e), "quote": ""}


# --- MAIN EXECUTION ---
def main():
    # 1. Load Rules
    try:
        rubric_dict = checklist.extract_rubric_to_dict(CHECKLIST_PATH)
        print(f"Loaded {len(rubric_dict)} requirements from checklist.")
    except Exception as e:
        print(f"CRITICAL: Could not load checklist: {e}")
        return

    # 2. Setup
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(INDEX_DIR, exist_ok=True)

    # 3. Build/load tender-docs index
    print("\n--- Preparing Tender Docs Index ---")
    tender_index = build_or_load_index("tender-docs", TENDER_DOCS_ROOT)
    tender_engine = tender_index.as_query_engine(similarity_top_k=10)
    tender_tool = QueryEngineTool.from_defaults(
        query_engine=tender_engine,
        name="tender_context",
        description="Use this to understand what a requirement means based on the tender documents."
    )

    # 4. Prepare combined workbook
    out_path = os.path.join(OUTPUT_DIR, "Report_Agent.xlsx")
    wb = load_workbook(CHECKLIST_PATH)
    template_ws = wb.active
    template_ws.title = "_template"

    # 5. Iterate bidder folders
    evidence_root = Path(EVIDENCE_ROOT)
    for folder in [f for f in evidence_root.iterdir() if f.is_dir()]:
        print(f"\n--- Processing: {folder.name} ---")

        # A. Build/load bidder index
        try:
            bidder_index = build_or_load_index(folder.name, str(folder))
        except Exception as e:
            print(f"   [Error] Could not build index for {folder.name}: {e}")
            continue

        bidder_engine = bidder_index.as_query_engine(similarity_top_k=10)
        bidder_tool = QueryEngineTool.from_defaults(
            query_engine=bidder_engine,
            name="bidder_docs",
            description="Use this to find evidence in the bidder's submitted documents."
        )

        # B. Create ReActAgent with both tools
        agent = ReActAgent.from_tools(
            [tender_tool, bidder_tool],
            llm=Settings.llm,
            verbose=True,
            max_iterations=6,
        )

        # C. Add sheet for this bidder
        ws = wb.copy_worksheet(template_ws)
        ws.title = folder.name[:31]

        print(f"   Auditing requirements...", end="", flush=True)

        # D. Audit each requirement
        for row in ws.iter_rows(min_row=2):
            cell_id_val = row[COL_ID - 1].value

            if cell_id_val:
                req_id = str(cell_id_val)

                if req_id in rubric_dict:
                    req_text = rubric_dict[req_id]

                    verdict = audit_with_agent(agent, req_id, req_text)

                    row[COL_STATUS - 1].value = verdict.get("status", "Error")
                    row[COL_EVIDENCE - 1].value = f"{verdict.get('reasoning', '')}\n\nRef: \"{verdict.get('quote', '')}\""

                    print(".", end="", flush=True)

        print(f"\n   Sheet added: {folder.name}")

    # 6. Remove template sheet and save
    wb.remove(template_ws)
    wb.save(out_path)
    print(f"\nAgent report saved: {out_path}")


if __name__ == "__main__":
    main()
